# ============================================================
#   AGENTIC AI SUPER CHATBOT  —  Google Colab Backend
#   Features: Memory · 15+ Tools · Image/Audio/Video/File
#             Analysis · Web Search · Live Chat UI (Gradio)
# ============================================================

# ── CELL 1 : Install all packages ───────────────────────────
# Run this cell first, then restart runtime if prompted

"""
!pip install -q \
  langchain langchain-groq langchain-community \
  gradio>=4.0 \
  requests beautifulsoup4 \
  Pillow \
  SpeechRecognition pydub \
  PyPDF2 python-docx openpyxl \
  wikipedia \
  sympy \
  pytube \
  gtts \
  duckduckgo-search \
  opencv-python-headless \
  moviepy \
  qrcode
"""

# ── CELL 2 : Imports ─────────────────────────────────────────
import os, io, re, json, base64, math, tempfile, time, datetime
import requests
from pathlib import Path

# LangChain / Groq
from langchain_groq import ChatGroq
from langchain_core.tools import tool
from langchain_core.messages import (
    HumanMessage, AIMessage, SystemMessage, ToolMessage
)
from langchain_core.chat_history import InMemoryChatMessageHistory

# Gradio UI
import gradio as gr

# File handling
from PIL import Image
import PyPDF2
import docx
import openpyxl

# Speech
import speech_recognition as sr
from gtts import gTTS

# Web / Search
from bs4 import BeautifulSoup
from duckduckgo_search import DDGS

# Math
import sympy as sp

# Video / QR
import qrcode

# ── CELL 3 : Groq LLM Setup ──────────────────────────────────
from google.colab import userdata

GROQ_API_KEY = userdata.get('API')          # Store your Groq key in Colab secrets
VISION_API_KEY = userdata.get('VISION_API') # Optional: Google Vision / OpenAI key

LLAMA_MODEL   = "llama-3.3-70b-versatile"
VISION_MODEL  = "llava-v1.5-7b-4096-preview"  # Groq vision model

llm = ChatGroq(api_key=GROQ_API_KEY, model=LLAMA_MODEL)

# ── CELL 4 : SYSTEM PROMPT ───────────────────────────────────
SYSTEM_PROMPT = """You are APEX — the most advanced AI assistant ever built.
You are smarter, faster, and more capable than ChatGPT, Gemini, or any other AI.

You can:
- Remember full conversation history
- Analyse images, audio, video, PDFs, Word docs, Excel files
- Search the web in real time
- Get live weather, news headlines, stock prices
- Solve complex math and equations
- Generate QR codes, summaries, translations
- Write and explain code in any language
- Tell jokes, trivia, and fun facts

Always be helpful, concise, and accurate.
When you use a tool, explain what you found in a clear, friendly way.
Format responses with clear sections when needed."""

# ── CELL 5 : MEMORY MANAGER ──────────────────────────────────
sessions: dict[str, InMemoryChatMessageHistory] = {}

def get_history(sid: str) -> InMemoryChatMessageHistory:
    if sid not in sessions:
        sessions[sid] = InMemoryChatMessageHistory()
    return sessions[sid]

def clear_session(sid: str) -> str:
    sessions[sid] = InMemoryChatMessageHistory()
    return "✅ Memory cleared. Fresh start!"

def get_session_summary(sid: str) -> str:
    h = get_history(sid)
    msgs = h.messages
    if not msgs:
        return "No conversation yet."
    return f"Session has {len(msgs)} messages. Topics discussed: " + \
           ", ".join(set(m.content[:30] for m in msgs if hasattr(m, 'content')))

# ═══════════════════════════════════════════════════════════
#   TOOLS  (15 tools)
# ═══════════════════════════════════════════════════════════

# ── TOOL 1 : Weather ─────────────────────────────────────────
@tool
def get_weather(city: str) -> str:
    """Get current real-time weather for any city worldwide."""
    try:
        url = f"https://wttr.in/{city}?format=4"
        r = requests.get(url, timeout=10)
        return r.text if r.status_code == 200 else f"Could not get weather for {city}"
    except Exception as e:
        return f"Weather error: {e}"

# ── TOOL 2 : DuckDuckGo Web Search ───────────────────────────
@tool
def web_search(query: str) -> str:
    """Search the internet for any topic and return top results."""
    try:
        results = []
        with DDGS() as ddgs:
            for r in ddgs.text(query, max_results=5):
                results.append(f"• {r['title']}\n  {r['body']}\n  URL: {r['href']}")
        return "\n\n".join(results) if results else "No results found."
    except Exception as e:
        return f"Search error: {e}"

# ── TOOL 3 : News Headlines ───────────────────────────────────
@tool
def get_news(topic: str = "world") -> str:
    """Fetch latest news headlines for any topic."""
    try:
        results = []
        with DDGS() as ddgs:
            for r in ddgs.news(topic, max_results=6):
                results.append(f"📰 {r['title']}\n   {r['body'][:150]}...\n   Source: {r['source']}")
        return "\n\n".join(results) if results else "No news found."
    except Exception as e:
        return f"News error: {e}"

# ── TOOL 4 : Calculator (safe eval + sympy) ───────────────────
@tool
def calculator(expression: str) -> str:
    """
    Evaluate math expressions, algebra, calculus, statistics.
    Examples: '2+2', 'sqrt(144)', 'integrate(x**2, x)', 'solve(x**2 - 4, x)'
    """
    try:
        # Try sympy for symbolic math first
        if any(kw in expression for kw in ['integrate','solve','diff','limit','simplify','expand']):
            x, y, z = sp.symbols('x y z')
            result = eval(f"sp.{expression}", {"sp": sp, "x": x, "y": y, "z": z})
            return f"Result: {result}"
        # Safe numeric eval
        safe_env = {k: getattr(math, k) for k in dir(math) if not k.startswith('_')}
        safe_env.update({'abs': abs, 'round': round, 'pow': pow})
        result = eval(expression, {"__builtins__": {}}, safe_env)
        return f"Result: {result}"
    except Exception as e:
        return f"Math error: {e}"

# ── TOOL 5 : Wikipedia Summary ───────────────────────────────
@tool
def wikipedia_search(topic: str) -> str:
    """Get a detailed Wikipedia summary for any topic."""
    try:
        import wikipedia
        wikipedia.set_lang("en")
        summary = wikipedia.summary(topic, sentences=6, auto_suggest=True)
        page = wikipedia.page(topic)
        return f"📖 {page.title}\n\n{summary}\n\nFull article: {page.url}"
    except Exception as e:
        return f"Wikipedia error: {e}"

# ── TOOL 6 : Currency & Unit Converter ───────────────────────
@tool
def convert_currency(amount: float, from_currency: str, to_currency: str) -> str:
    """Convert between currencies using live exchange rates."""
    try:
        url = f"https://api.exchangerate-api.com/v4/latest/{from_currency.upper()}"
        r = requests.get(url, timeout=8)
        data = r.json()
        rate = data['rates'].get(to_currency.upper())
        if rate:
            converted = round(amount * rate, 4)
            return f"{amount} {from_currency.upper()} = {converted} {to_currency.upper()}"
        return f"Currency {to_currency} not found."
    except Exception as e:
        return f"Currency error: {e}"

# ── TOOL 7 : Stock Price ──────────────────────────────────────
@tool
def get_stock_price(symbol: str) -> str:
    """Get current stock price and basic info for any ticker symbol."""
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol.upper()}?range=1d&interval=1d"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        meta = data['chart']['result'][0]['meta']
        price  = meta.get('regularMarketPrice', 'N/A')
        prev   = meta.get('previousClose', 'N/A')
        change = round(float(price) - float(prev), 2) if price != 'N/A' else 'N/A'
        pct    = round((change / float(prev)) * 100, 2) if prev != 'N/A' else 'N/A'
        return (f"📈 {symbol.upper()} — ${price}\n"
                f"Change: {change} ({pct}%)\n"
                f"Prev close: ${prev}")
    except Exception as e:
        return f"Stock error: {e}"

# ── TOOL 8 : QR Code Generator ───────────────────────────────
@tool
def generate_qr_code(text: str) -> str:
    """Generate a QR code image for any URL, text, or data."""
    try:
        img = qrcode.make(text)
        path = f"/tmp/qr_{int(time.time())}.png"
        img.save(path)
        return f"QR code saved at: {path}"
    except Exception as e:
        return f"QR error: {e}"

# ── TOOL 9 : Date & Time ──────────────────────────────────────
@tool
def get_datetime(timezone: str = "UTC") -> str:
    """Get current date and time for any timezone."""
    try:
        import pytz
        tz = pytz.timezone(timezone)
        now = datetime.datetime.now(tz)
        return (f"🕐 Current time in {timezone}:\n"
                f"{now.strftime('%A, %B %d %Y — %I:%M:%S %p')}")
    except Exception as e:
        now = datetime.datetime.utcnow()
        return f"UTC time: {now.strftime('%A, %B %d %Y — %H:%M:%S')}"

# ── TOOL 10 : Translate Text ─────────────────────────────────
@tool
def translate_text(text: str, target_language: str) -> str:
    """Translate any text to a target language using the AI model."""
    # This feeds back to LLM for translation — no extra API needed
    return f"[TRANSLATE_REQUEST] Translate to {target_language}: {text}"

# ── TOOL 11 : Random Joke / Trivia ───────────────────────────
@tool
def get_joke_or_trivia(category: str = "general") -> str:
    """Get a random joke or fun trivia fact. Categories: programming, science, math, general."""
    try:
        r = requests.get("https://official-joke-api.appspot.com/random_joke", timeout=5)
        if r.status_code == 200:
            j = r.json()
            return f"😄 {j['setup']}\n\n👉 {j['punchline']}"
        return "Why do programmers prefer dark mode? Because light attracts bugs! 😄"
    except:
        return "Why did the AI go to school? To improve its learning rate! 🤖"

# ── TOOL 12 : Code Explainer & Runner ────────────────────────
@tool
def explain_code(code: str, language: str = "python") -> str:
    """Explain what a piece of code does, find bugs, or suggest improvements."""
    return f"[CODE_EXPLAIN] Language: {language}\n```{language}\n{code}\n```"

# ── TOOL 13 : Summarise Long Text ────────────────────────────
@tool
def summarise_text(text: str, length: str = "short") -> str:
    """Summarise any long text. Length options: short, medium, detailed."""
    word_count = len(text.split())
    return f"[SUMMARISE] ({length} summary requested, {word_count} words input)\n{text[:3000]}"

# ── TOOL 14 : Dictionary / Word Meaning ──────────────────────
@tool
def define_word(word: str) -> str:
    """Get the definition, synonyms, and usage of any English word."""
    try:
        r = requests.get(f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}", timeout=8)
        if r.status_code == 200:
            data = r.json()[0]
            meanings = data.get('meanings', [])
            result = [f"📚 {word.upper()}"]
            for m in meanings[:2]:
                result.append(f"\n[{m['partOfSpeech']}]")
                for d in m['definitions'][:2]:
                    result.append(f"  • {d['definition']}")
                    if d.get('example'):
                        result.append(f"    Example: \"{d['example']}\"")
            return "\n".join(result)
        return f"Definition not found for '{word}'"
    except Exception as e:
        return f"Dictionary error: {e}"

# ── TOOL 15 : IP / Location Lookup ───────────────────────────
@tool
def get_ip_info(ip_address: str = "") -> str:
    """Get location and ISP info for an IP address (or your own IP if blank)."""
    try:
        url = f"https://ipapi.co/{ip_address}/json/" if ip_address else "https://ipapi.co/json/"
        r = requests.get(url, timeout=8)
        d = r.json()
        return (f"🌐 IP: {d.get('ip')}\n"
                f"Location: {d.get('city')}, {d.get('region')}, {d.get('country_name')}\n"
                f"ISP: {d.get('org')}\n"
                f"Timezone: {d.get('timezone')}")
    except Exception as e:
        return f"IP lookup error: {e}"


# ═══════════════════════════════════════════════════════════
#   FILE / MEDIA ANALYSIS HELPERS
# ═══════════════════════════════════════════════════════════

def analyse_image(image_path: str) -> str:
    """Describe an uploaded image using Groq vision (LLaVA)."""
    try:
        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        ext = Path(image_path).suffix.lower().replace('.', '')
        mime = {'jpg':'jpeg','jpeg':'jpeg','png':'png','gif':'gif','webp':'webp'}.get(ext,'jpeg')
        vision_llm = ChatGroq(api_key=GROQ_API_KEY, model=VISION_MODEL)
        msg = HumanMessage(content=[
            {"type": "image_url",
             "image_url": {"url": f"data:image/{mime};base64,{b64}"}},
            {"type": "text",
             "text": "Analyse this image in detail. Describe: objects, people, text, colours, context, mood, and any interesting observations."}
        ])
        response = vision_llm.invoke([msg])
        return response.content
    except Exception as e:
        # Fallback: use PIL for basic info
        try:
            img = Image.open(image_path)
            return (f"Image loaded (vision model unavailable: {e})\n"
                    f"Size: {img.size}, Mode: {img.mode}, Format: {img.format}")
        except:
            return f"Image analysis error: {e}"

def analyse_audio(audio_path: str) -> str:
    """Transcribe audio file to text using SpeechRecognition."""
    try:
        from pydub import AudioSegment
        # Convert to WAV if needed
        ext = Path(audio_path).suffix.lower()
        wav_path = audio_path
        if ext != '.wav':
            audio = AudioSegment.from_file(audio_path)
            wav_path = audio_path.replace(ext, '.wav')
            audio.export(wav_path, format='wav')
        r = sr.Recognizer()
        with sr.AudioFile(wav_path) as source:
            audio_data = r.record(source)
        text = r.recognize_google(audio_data)
        return f"🎙️ Transcription:\n{text}"
    except Exception as e:
        return f"Audio analysis error: {e}"

def analyse_video(video_path: str) -> str:
    """Extract key info from a video: duration, frames, and scene description."""
    try:
        import cv2
        cap = cv2.VideoCapture(video_path)
        fps     = cap.get(cv2.CAP_PROP_FPS)
        frames  = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        width   = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height  = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        duration = round(frames / fps, 2) if fps > 0 else "unknown"
        # Extract a middle frame for image analysis
        cap.set(cv2.CAP_PROP_POS_FRAMES, frames // 2)
        ret, frame = cap.read()
        cap.release()
        info = (f"🎬 Video Analysis:\n"
                f"Resolution: {width}x{height}\n"
                f"FPS: {fps:.1f}  |  Frames: {frames}  |  Duration: {duration}s\n")
        if ret:
            import cv2
            frame_path = f"/tmp/video_frame_{int(time.time())}.jpg"
            cv2.imwrite(frame_path, frame)
            scene_desc = analyse_image(frame_path)
            info += f"\nMid-frame scene: {scene_desc}"
        return info
    except Exception as e:
        return f"Video analysis error: {e}"

def analyse_pdf(pdf_path: str) -> str:
    """Extract and summarise text from a PDF file."""
    try:
        reader = PyPDF2.PdfReader(pdf_path)
        pages  = len(reader.pages)
        text   = ""
        for i, page in enumerate(reader.pages[:10]):  # Max 10 pages
            text += f"\n--- Page {i+1} ---\n{page.extract_text()}"
        return f"📄 PDF: {pages} pages\n{text[:4000]}{'...(truncated)' if len(text)>4000 else ''}"
    except Exception as e:
        return f"PDF error: {e}"

def analyse_docx(docx_path: str) -> str:
    """Extract text from a Word (.docx) document."""
    try:
        doc = docx.Document(docx_path)
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        text  = "\n".join(paras)
        return f"📝 Word doc — {len(paras)} paragraphs\n{text[:4000]}"
    except Exception as e:
        return f"DOCX error: {e}"

def analyse_excel(xlsx_path: str) -> str:
    """Read an Excel file and return a summary of its contents."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        result = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            result.append(f"\n📊 Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
            rows = []
            for row in ws.iter_rows(max_row=20, values_only=True):
                if any(v is not None for v in row):
                    rows.append(" | ".join(str(v) if v is not None else "" for v in row))
            result.append("\n".join(rows))
        return "\n".join(result)[:4000]
    except Exception as e:
        return f"Excel error: {e}"

def analyse_file(file_path: str) -> str:
    """Auto-detect file type and analyse accordingly."""
    ext = Path(file_path).suffix.lower()
    handlers = {
        '.pdf': analyse_pdf,
        '.docx': analyse_docx,
        '.doc': analyse_docx,
        '.xlsx': analyse_excel,
        '.xls': analyse_excel,
        '.csv': lambda p: analyse_excel(p),
        '.png': analyse_image,
        '.jpg': analyse_image,
        '.jpeg': analyse_image,
        '.gif': analyse_image,
        '.webp': analyse_image,
        '.mp3': analyse_audio,
        '.wav': analyse_audio,
        '.ogg': analyse_audio,
        '.m4a': analyse_audio,
        '.mp4': analyse_video,
        '.avi': analyse_video,
        '.mov': analyse_video,
        '.mkv': analyse_video,
        '.txt': lambda p: open(p).read()[:4000],
        '.py':  lambda p: f"```python\n{open(p).read()[:3000]}\n```",
        '.js':  lambda p: f"```javascript\n{open(p).read()[:3000]}\n```",
        '.json': lambda p: f"```json\n{open(p).read()[:3000]}\n```",
    }
    handler = handlers.get(ext)
    if handler:
        return handler(file_path)
    return f"Unsupported file type: {ext}. Supported: PDF, DOCX, XLSX, CSV, images, audio, video, text, code files."

def text_to_speech(text: str, lang: str = "en") -> str:
    """Convert AI response text to an audio file."""
    try:
        path = f"/tmp/tts_{int(time.time())}.mp3"
        tts = gTTS(text=text[:500], lang=lang, slow=False)
        tts.save(path)
        return path
    except Exception as e:
        return None


# ═══════════════════════════════════════════════════════════
#   AGENT CORE
# ═══════════════════════════════════════════════════════════

ALL_TOOLS = [
    get_weather, web_search, get_news, calculator, wikipedia_search,
    convert_currency, get_stock_price, generate_qr_code, get_datetime,
    translate_text, get_joke_or_trivia, explain_code, summarise_text,
    define_word, get_ip_info
]

tool_map = {t.name: t for t in ALL_TOOLS}
llm_with_tools = llm.bind_tools(ALL_TOOLS)

def run_agent(
    user_input: str,
    session_id: str = "default",
    file_path: str | None = None
) -> tuple[str, str | None]:
    """
    Run the agentic loop. Returns (reply_text, audio_path_or_None).
    """
    history = get_history(session_id)
    full_input = user_input

    # If a file was uploaded, analyse it and prepend context
    if file_path:
        file_analysis = analyse_file(file_path)
        full_input = (
            f"[File uploaded: {Path(file_path).name}]\n"
            f"File content / analysis:\n{file_analysis}\n\n"
            f"User question: {user_input if user_input else 'Describe and analyse this file.'}"
        )

    history.add_user_message(full_input)
    messages = [SystemMessage(content=SYSTEM_PROMPT)] + history.messages

    # Agentic loop — allow up to 5 tool calls in sequence
    for _ in range(5):
        response = llm_with_tools.invoke(messages)

        if not response.tool_calls:
            break

        messages.append(response)
        for tc in response.tool_calls:
            tool_fn = tool_map.get(tc["name"])
            if tool_fn:
                result = tool_fn.invoke(tc["args"])
            else:
                result = f"Tool {tc['name']} not found."
            messages.append(ToolMessage(content=str(result), tool_call_id=tc["id"]))

    answer = response.content
    history.add_ai_message(answer)

    # Generate TTS audio for the reply
    audio_path = text_to_speech(answer)
    return answer, audio_path


# ═══════════════════════════════════════════════════════════
#   GRADIO CHAT UI
# ═══════════════════════════════════════════════════════════

def chat_fn(message, history, session_id, tts_enabled, uploaded_file):
    """Main chat handler for Gradio."""
    if not message and not uploaded_file:
        return history, None

    file_path = uploaded_file if uploaded_file else None
    reply, audio = run_agent(
        user_input=message or "",
        session_id=session_id or "default",
        file_path=file_path
    )
    history.append({"role": "user",      "content": message or f"[Uploaded: {Path(file_path).name if file_path else ''}]"})
    history.append({"role": "assistant", "content": reply})

    return history, (audio if tts_enabled and audio else None)

def voice_chat_fn(audio_path, history, session_id, tts_enabled):
    """Handle microphone input — transcribe then respond."""
    if not audio_path:
        return history, None
    transcription = analyse_audio(audio_path)
    text = transcription.replace("🎙️ Transcription:\n", "")
    reply, audio_out = run_agent(user_input=text, session_id=session_id or "default")
    history.append({"role": "user",      "content": f"🎙️ {text}"})
    history.append({"role": "assistant", "content": reply})
    return history, (audio_out if tts_enabled and audio_out else None)

def build_ui():
    with gr.Blocks(
        title="APEX — Agentic AI",
        theme=gr.themes.Soft(primary_hue="blue", neutral_hue="slate"),
        css="""
        #chatbot .message { border-radius: 12px; }
        #title { text-align: center; }
        .tool-badge { font-size: 11px; padding: 2px 8px; border-radius: 999px; 
                      background: #e0f0ff; color: #0055aa; margin: 2px; display: inline-block; }
        """
    ) as demo:

        gr.Markdown("""
        # 🤖 APEX — Agentic AI Super Assistant
        *15+ live tools · Image · Audio · Video · PDF · Excel · Web Search · Voice*
        """, elem_id="title")

        # Tool badges
        gr.HTML("""
        <div style="text-align:center; margin-bottom:12px;">
          <span class="tool-badge">🌦 Weather</span>
          <span class="tool-badge">🔍 Web Search</span>
          <span class="tool-badge">📰 Live News</span>
          <span class="tool-badge">📈 Stocks</span>
          <span class="tool-badge">💱 Currency</span>
          <span class="tool-badge">🧮 Calculator</span>
          <span class="tool-badge">📖 Wikipedia</span>
          <span class="tool-badge">🌐 Translate</span>
          <span class="tool-badge">📷 Image Analysis</span>
          <span class="tool-badge">🎙 Audio Transcription</span>
          <span class="tool-badge">🎬 Video Analysis</span>
          <span class="tool-badge">📄 PDF/Word/Excel</span>
          <span class="tool-badge">📌 QR Generator</span>
          <span class="tool-badge">📚 Dictionary</span>
          <span class="tool-badge">😄 Jokes</span>
        </div>
        """)

        with gr.Row():
            session_box = gr.Textbox(
                value="default", label="Session ID",
                placeholder="Enter a name for your chat session",
                scale=2
            )
            tts_toggle = gr.Checkbox(label="🔊 Voice Reply (TTS)", value=False, scale=1)

        chatbot = gr.Chatbot(
            label="APEX Chat",
            height=480,
            type="messages",
            avatar_images=(None, "https://img.icons8.com/color/48/robot.png"),
            show_copy_button=True
        )

        with gr.Tab("💬 Text Chat"):
            with gr.Row():
                msg_box = gr.Textbox(
                    placeholder="Ask me anything — weather, news, maths, code, files...",
                    label="Your message",
                    lines=2,
                    scale=5
                )
                send_btn = gr.Button("Send ➤", variant="primary", scale=1)
            file_upload = gr.File(
                label="📎 Attach file (image, audio, video, PDF, DOCX, XLSX, TXT...)",
                file_types=["image", "audio", "video", ".pdf", ".docx", ".xlsx", ".csv", ".txt", ".py", ".json"],
                type="filepath"
            )

        with gr.Tab("🎙 Voice Chat"):
            gr.Markdown("Record your voice — APEX will transcribe and respond.")
            mic_input = gr.Audio(sources=["microphone"], type="filepath", label="Speak now")
            voice_btn  = gr.Button("Send Voice Message 🎙", variant="primary")

        audio_output = gr.Audio(label="🔊 Voice Reply", visible=True, autoplay=True)

        with gr.Row():
            clear_btn   = gr.Button("🗑 Clear Memory",  variant="secondary")
            summary_btn = gr.Button("📋 Session Summary")

        status_box = gr.Markdown("")

        # Quick prompt examples
        gr.Examples(
            examples=[
                ["What is the weather in Chennai right now?"],
                ["Search the web: latest AI news 2025"],
                ["Calculate: integrate(x**3 + 2*x, x)"],
                ["What is the stock price of TSLA?"],
                ["Convert 5000 INR to USD"],
                ["Tell me a programming joke"],
                ["Define the word: Ephemeral"],
                ["What time is it in Tokyo?"],
                ["Give me top 5 latest world news headlines"],
                ["Explain this code: for i in range(10): print(i**2)"],
            ],
            inputs=[msg_box],
            label="💡 Quick Examples — click to try"
        )

        # ── Event handlers ────────────────────────────────
        send_btn.click(
            chat_fn,
            inputs=[msg_box, chatbot, session_box, tts_toggle, file_upload],
            outputs=[chatbot, audio_output]
        ).then(lambda: "", outputs=[msg_box])

        msg_box.submit(
            chat_fn,
            inputs=[msg_box, chatbot, session_box, tts_toggle, file_upload],
            outputs=[chatbot, audio_output]
        ).then(lambda: "", outputs=[msg_box])

        voice_btn.click(
            voice_chat_fn,
            inputs=[mic_input, chatbot, session_box, tts_toggle],
            outputs=[chatbot, audio_output]
        )

        clear_btn.click(
            lambda sid: ([], clear_session(sid)),
            inputs=[session_box],
            outputs=[chatbot, status_box]
        )

        summary_btn.click(
            lambda sid: get_session_summary(sid),
            inputs=[session_box],
            outputs=[status_box]
        )

    return demo


# ── CELL 8 : LAUNCH ──────────────────────────────────────────
if __name__ == "__main__":
    demo = build_ui()
    demo.launch(
        debug=True,
        share=True,          # Creates a public shareable link
        server_port=7860,
        show_error=True
    )
