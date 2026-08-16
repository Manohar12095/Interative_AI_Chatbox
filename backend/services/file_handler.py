"""
File analysis handlers — ported from agentic_chatbot.py
"""
import base64
import time
from pathlib import Path
from PIL import Image
import PyPDF2
import docx
import openpyxl
from config import GROQ_API_KEY, VISION_MODEL, UPLOADS_DIR


def analyse_image(image_path: str) -> str:
    """Describe an uploaded image using Groq vision (qwen/qwen3.6-27b) with PIL fallback."""
    import logging
    logger = logging.getLogger(__name__)
    try:
        from langchain_groq import ChatGroq
        from langchain_core.messages import HumanMessage

        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        ext = Path(image_path).suffix.lower().replace('.', '')
        mime = {'jpg': 'jpeg', 'jpeg': 'jpeg', 'png': 'png', 'gif': 'gif', 'webp': 'webp'}.get(ext, 'jpeg')

        logger.info(f"Analysing image with vision model: {VISION_MODEL}")
        vision_llm = ChatGroq(api_key=GROQ_API_KEY, model=VISION_MODEL)
        msg = HumanMessage(content=[
            {"type": "image_url", "image_url": {"url": f"data:image/{mime};base64,{b64}"}},
            {"type": "text", "text": "Analyse this image in detail. Describe what you see: objects, people, text, colours, spatial relationships, context, and overall mood."}
        ])
        response = vision_llm.invoke([msg])
        return response.content
    except Exception as llm_err:
        # Log the real error server-side so it can be diagnosed
        logger.error(f"Vision LLM call failed (model={VISION_MODEL}): {llm_err}")
        # Graceful degradation — try to at least return PIL metadata
        try:
            img = Image.open(image_path)
            return (
                f"⚠️ Vision analysis temporarily unavailable (model error logged server-side).\n"
                f"Image info — Size: {img.size[0]}×{img.size[1]}px, Mode: {img.mode}, Format: {img.format}"
            )
        except Exception as pil_err:
            logger.error(f"PIL fallback also failed: {pil_err}")
            return f"Image analysis error: {llm_err}"


def analyse_audio(audio_path: str) -> str:
    """Transcribe audio file to text."""
    try:
        import speech_recognition as sr
        from pydub import AudioSegment

        ext = Path(audio_path).suffix.lower()
        wav_path = audio_path
        if ext != '.wav':
            audio = AudioSegment.from_file(audio_path)
            wav_path = audio_path.replace(ext, '.wav')
            audio.export(wav_path, format='wav')

        r = sr.Recognizer()
        with sr.AudioFile(wav_path) as source:
            audio_data = r.record(source)
        text = r.recognize_google(audio_data)
        return f"🎙️ Transcription:\n{text}"
    except Exception as e:
        return f"Audio analysis error: {e}"


def analyse_pdf(pdf_path: str) -> str:
    """Extract text from a PDF file."""
    try:
        reader = PyPDF2.PdfReader(pdf_path)
        pages = len(reader.pages)
        text = ""
        for i, page in enumerate(reader.pages[:10]):
            text += f"\n--- Page {i + 1} ---\n{page.extract_text()}"
        return f"📄 PDF: {pages} pages\n{text[:4000]}{'...(truncated)' if len(text) > 4000 else ''}"
    except Exception as e:
        return f"PDF error: {e}"


def analyse_docx(docx_path: str) -> str:
    """Extract text from a Word (.docx) document."""
    try:
        doc = docx.Document(docx_path)
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paras)
        return f"📝 Word doc — {len(paras)} paragraphs\n{text[:4000]}"
    except Exception as e:
        return f"DOCX error: {e}"


def analyse_excel(xlsx_path: str) -> str:
    """Read an Excel file and return a summary."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        result = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            result.append(f"\n📊 Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
            rows = []
            for row in ws.iter_rows(max_row=20, values_only=True):
                if any(v is not None for v in row):
                    rows.append(" | ".join(str(v) if v is not None else "" for v in row))
            result.append("\n".join(rows))
        return "\n".join(result)[:4000]
    except Exception as e:
        return f"Excel error: {e}"


def analyse_file(file_path: str) -> str:
    """Auto-detect file type and analyse accordingly."""
    ext = Path(file_path).suffix.lower()
    handlers = {
        '.pdf': analyse_pdf,
        '.docx': analyse_docx,
        '.doc': analyse_docx,
        '.xlsx': analyse_excel,
        '.xls': analyse_excel,
        '.csv': lambda p: analyse_excel(p),
        '.png': analyse_image,
        '.jpg': analyse_image,
        '.jpeg': analyse_image,
        '.gif': analyse_image,
        '.webp': analyse_image,
        '.mp3': analyse_audio,
        '.wav': analyse_audio,
        '.ogg': analyse_audio,
        '.m4a': analyse_audio,
        '.txt': lambda p: open(p, encoding='utf-8', errors='ignore').read()[:4000],
        '.py': lambda p: f"```python\n{open(p, encoding='utf-8', errors='ignore').read()[:3000]}\n```",
        '.js': lambda p: f"```javascript\n{open(p, encoding='utf-8', errors='ignore').read()[:3000]}\n```",
        '.json': lambda p: f"```json\n{open(p, encoding='utf-8', errors='ignore').read()[:3000]}\n```",
    }
    handler = handlers.get(ext)
    if handler:
        return handler(file_path)
    return f"Unsupported file type: {ext}"
